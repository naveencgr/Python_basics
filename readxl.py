import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("music.xlsx")

sheet = wb["music"]
print(sheet.max_row)
print(sheet.max_column)
for i in range(1, sheet.max_row + 1):
    myList = []
    for j in range(1, sheet.max_column + 1):
        myList.append(sheet.cell(i, j).value) 
    print(myList) 

for i in range(1, sheet.max_row + 1):
    cell_update = sheet.cell(i, 4)
    cell_value = sheet.cell(i, 1).value
    cell_update.value =  cell_value * 10   


values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col = 4, max_col = 4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "e2")


wb.save("music_update.xlsx")       

